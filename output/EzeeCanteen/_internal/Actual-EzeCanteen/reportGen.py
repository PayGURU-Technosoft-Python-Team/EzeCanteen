import os
import pandas as pd
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Database configuration
DB_HOST = "103.216.211.36"
DB_USER = "pgcanteen"
DB_PORT = 33975
DB_PASS = "L^{Z,8~zzfF9(nd8"
DB_NAME = "payguru_canteen"
DB_TABLE = "sequentiallog"

def connect_to_database():
    """
    Create and return a connection to the database
    """
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            port=DB_PORT,
            password=DB_PASS,
            database=DB_NAME
        )
        return conn
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

def generate_monthly_report(year, month, output_dir="Reports/monthly"):
    """
    Generate a monthly report showing count of meal types per day

    Args:
        year (int): Year for the report
        month (int): Month for the report (1-12)
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate inputs
        if not isinstance(year, int) or not isinstance(month, int):
            raise ValueError("Year and month must be integers")
        
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Get distinct meal types
        meal_types_query = f"""
            SELECT DISTINCT Fooditem 
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        """
        cursor.execute(meal_types_query, (year, month))
        meal_types = [row['Fooditem'] for row in cursor.fetchall()]
        
        # Query to get meal counts by date for the specified month
        query = f"""
            SELECT 
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                Fooditem,
                COUNT(*) as Count
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
            GROUP BY DATE_FORMAT(PunchDateTime, '%d-%m-%Y'), Fooditem
            ORDER BY DATE(PunchDateTime)
        """
        
        cursor.execute(query, (year, month))
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {month}/{year}")
            # Create empty dataframe with dynamic columns
            columns = ["Date"] + meal_types + ["Total"]
            df = pd.DataFrame(columns=columns)
        else:
            # Convert to DataFrame for easier manipulation
            data_df = pd.DataFrame(results)
            
            # Create a pivot table
            pivot_df = pd.pivot_table(
                data_df,
                values="Count",
                index="PunchDate",
                columns="Fooditem",
                aggfunc='sum',
                fill_value=0
            )
            
            # Ensure all meal types from query are present
            for meal_type in meal_types:
                if meal_type not in pivot_df.columns:
                    pivot_df[meal_type] = 0
            
            # Keep only the required columns and reorder them
            pivot_df = pivot_df[meal_types]
            
            # Add Total column
            pivot_df["Total"] = pivot_df.sum(axis=1)
            
            # Reset index to make PunchDate a column
            pivot_df = pivot_df.reset_index()
            
            # Rename column to match expected format
            pivot_df = pivot_df.rename(columns={"PunchDate": "Date"})
            
            # Calculate totals
            totals = pivot_df[meal_types + ["Total"]].sum().to_frame().T
            totals.insert(0, "Date", "Total")
            
            # Combine dataframe with totals
            df = pd.concat([pivot_df, totals], ignore_index=True)
        
        # Create Excel file
        file_name = f"{year}_{month:02d}_CanteenReports.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Monthly Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Monthly Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align except for the first column
                    if col_num > 1:
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Format total row with bold
                    if row_num == len(df) + 1:
                        cell.font = Font(bold=True)
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Monthly report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating monthly report: {e}")
        return None

def generate_daily_report(date, output_dir="Reports/daily"):
    """
    Generate a daily report showing all entries for a specific date
    
    Args:
        date (str): Date in format YYYY-MM-DD
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate date format
        datetime.strptime(date, "%Y-%m-%d")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to get all entries for the specified date
        query = f"""
            SELECT 
                PunchCardNo,
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                TIME_FORMAT(PunchDateTime, '%H:%i:%s') as PunchTime,
                Fooditem
            FROM {DB_TABLE}
            WHERE DATE(PunchDateTime) = '{date}'
            ORDER BY PunchDateTime
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {date}")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Punch ID",  "Punch Date", "Punch Time", "Meal Type"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Format the data
            df["Punch ID"] = df["PunchCardNo"]
            df["Meal Type"] = df["Fooditem"]
            
            # Select and reorder columns
            df = df[["Punch ID", "PunchDate", "PunchTime", "Meal Type"]]
            
            # Rename columns to match expected format
            df = df.rename(columns={"PunchDate": "Punch Date", "PunchTime": "Punch Time"})
        
        # Create Excel file
        formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%Y-%m-%d")
        file_name = f"{formatted_date}_DailyConsumption.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Daily Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Daily Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align time and meal type columns
                    if col_num in [4, 5]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Daily report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating daily report: {e}")
        return None

def generate_logs_report(year, month, output_dir="Reports/CanteenLogs"):
    """
    Generate a logs report showing all entries for a specific month
    
    Args:
        year (int): Year for the report
        month (int): Month for the report (1-12)
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate inputs
        if not isinstance(year, int) or not isinstance(month, int):
            raise ValueError("Year and month must be integers")
        
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to get all entries for the specified month
        query = f"""
            SELECT 
                PunchCardNo,
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                TIME_FORMAT(PunchDateTime, '%H:%i:%s') as PunchTime,
                Fooditem
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = {year} AND MONTH(PunchDateTime) = {month}
            ORDER BY PunchDateTime
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {month}/{year}")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Punch ID", "Punch Date", "Punch Time", "Meal Type"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Format the data
            # Since we can't join with configh for name, just use PunchCardNo
            df["Punch ID"] = df["PunchCardNo"]
            df["Meal Type"] = df["Fooditem"]
            
            # Select and reorder columns - use directly formatted columns from SQL
            df = df[["Punch ID", "PunchDate", "PunchTime", "Meal Type"]]
            
            # Rename columns to match expected format
            df = df.rename(columns={"PunchDate": "Punch Date", "PunchTime": "Punch Time"})
        
        # Create Excel file
        file_name = f"{year}_{month}_CanteenLogs.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Logs Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Logs Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align time and meal type columns
                    if col_num in [4, 5]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Logs report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating logs report: {e}")
        return None

def generate_fooditem_count_report(output_dir="Reports/fooditemEvaluation"):
    """
    Generate a report showing counts of each food item in the database
    
    Args:
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to count occurrences of each food item
        query = f"""
            SELECT 
                Fooditem,
                COUNT(*) as Count
            FROM {DB_TABLE}
            GROUP BY Fooditem
            ORDER BY COUNT(*) DESC
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print("No food items found in the database")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Food Item", "Count"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Rename columns to match expected format
            df = df.rename(columns={"Fooditem": "Food Item"})
        
        # Create Excel file
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_name = f"{current_date}_FoodItemCounts.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Food Item Counts', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Food Item Counts']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align count column
                    if col_num == 2:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Food item count report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating food item count report: {e}")
        return None

def generate_device_report(workbook, year, month, month_number, output_dir, prompt_for_location):
    """Generate a device-based monthly report with three worksheets"""
    try:
        # Create worksheets
        daily_worksheet = workbook.active
        daily_worksheet.title = "Daily Summary"
        user_worksheet = workbook.create_sheet("User Summary")
        detail_worksheet = workbook.create_sheet("Consumption Detail")
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all meal types for the month
        meal_types_query = f"""
            SELECT DISTINCT Fooditem 
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        """
        cursor.execute(meal_types_query, (year, month))
        meal_types = [row['Fooditem'] for row in cursor.fetchall()]
        
        # Fill Daily Summary worksheet
        columns = fill_device_daily_summary(daily_worksheet, cursor, year, month, meal_types)
        
        # Fill User Summary worksheet
        fill_device_user_summary(user_worksheet, cursor, year, month, meal_types)
        
        # Fill Consumption Detail worksheet
        column_headers = [col for col in columns if col != "Date" and col != "Total"]
        fill_device_consumption_detail(detail_worksheet, cursor, year, month, column_headers)
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Create the file path
        file_name = f"{year}_{month_number}_CanteenReports.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Ask for save location if prompted
        if prompt_for_location:
            # This would typically use a GUI dialog in a desktop app
            # For a script, we'll just use the output_dir
            pass
            
        # Save the workbook
        workbook.save(file_path)
        print(f"Device-based monthly report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating device report: {e}")
        import traceback
        traceback.print_exc()
        return None

def fill_device_daily_summary(worksheet, cursor, year, month, meal_types):
    """Fill the Daily Summary worksheet for device-based report"""
    # Define headers
    columns = ["Date"] + meal_types + ["Total"]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get daily counts by meal type
    query = f"""
        SELECT 
            DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
            Fooditem,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY DATE_FORMAT(PunchDateTime, '%d-%m-%Y'), Fooditem
        ORDER BY DATE(PunchDateTime)
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by date
    data_by_date = {}
    for row in results:
        date = row['PunchDate']
        meal = row['Fooditem']
        count = row['Count']
        
        if date not in data_by_date:
            data_by_date[date] = {meal_type: 0 for meal_type in meal_types}
            data_by_date[date]['Date'] = date
            data_by_date[date]['Total'] = 0
            
        data_by_date[date][meal] = count
        data_by_date[date]['Total'] += count
    
    # Add data rows
    row_idx = 2
    for date, data in sorted(data_by_date.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_idx > 1:  # Center align all but first column
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {meal_type: 0 for meal_type in meal_types}
    totals['Date'] = 'Total'
    totals['Total'] = 0
    
    for date_data in data_by_date.values():
        for meal_type in meal_types:
            totals[meal_type] += date_data.get(meal_type, 0)
        totals['Total'] += date_data.get('Total', 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, 0) 
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_idx > 1:  # Center align all but first column
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return columns

def fill_device_user_summary(worksheet, cursor, year, month, meal_types):
    """Fill the User Summary worksheet for device-based report"""
    # Define headers
    columns = ["Year", "Month", "PunchID", "Name"] + meal_types
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get user data
    query = f"""
        SELECT 
            PunchCardNo,
            Fooditem,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, Fooditem
        ORDER BY PunchCardNo
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user
    data_by_user = {}
    for row in results:
        user_id = row['PunchCardNo']
        meal = row['Fooditem']
        count = row['Count']
        
        if user_id not in data_by_user:
            data_by_user[user_id] = {meal_type: 0 for meal_type in meal_types}
            data_by_user[user_id]['Year'] = year
            data_by_user[user_id]['Month'] = month
            data_by_user[user_id]['PunchID'] = user_id
            data_by_user[user_id]['Name'] = ""  # No name available in this table
            
        data_by_user[user_id][meal] = count
    
    # Add data rows
    row_idx = 2
    for user_id, data in sorted(data_by_user.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0 if col_name in meal_types else "")
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_name in meal_types:  # Center align meal counts
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {meal_type: 0 for meal_type in meal_types}
    totals['PunchID'] = 'Total'
    
    for user_data in data_by_user.values():
        for meal_type in meal_types:
            totals[meal_type] += user_data.get(meal_type, 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, "")
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_name in meal_types:  # Center align meal counts
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def fill_device_consumption_detail(worksheet, cursor, year, month, meal_types):
    """Fill the Consumption Detail worksheet for device-based report"""
    # Define headers - Employee, Meal Label, Total, then day numbers 1-31
    columns = ["Employee", "Meal Label", "Total"] + [str(i) for i in range(1, 32)]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get detailed consumption data
    query = f"""
        SELECT 
            PunchCardNo,
            Fooditem,
            DAY(PunchDateTime) as Day,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, Fooditem, DAY(PunchDateTime)
        ORDER BY PunchCardNo, Fooditem, Day
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user and meal type
    data_by_user_meal = {}
    
    for row in results:
        user_id = row['PunchCardNo']
        meal_type = row['Fooditem']
        day = row['Day']
        count = row['Count']
        
        if user_id not in data_by_user_meal:
            data_by_user_meal[user_id] = {}
            
        if meal_type not in data_by_user_meal[user_id]:
            data_by_user_meal[user_id][meal_type] = {"Total": 0}
            
        data_by_user_meal[user_id][meal_type][str(day)] = count
        data_by_user_meal[user_id][meal_type]["Total"] += count
    
    # Add data rows
    row_idx = 2
    for user_id, meal_data in sorted(data_by_user_meal.items()):
        first_meal = True
        for meal_type, day_data in sorted(meal_data.items()):
            if meal_type not in meal_types:
                continue
                
            # First column: User ID (only for first meal type)
            if first_meal:
                cell = worksheet.cell(row=row_idx, column=1, value=f"{user_id}")
                cell.font = Font(bold=True)
            else:
                worksheet.cell(row=row_idx, column=1, value="")
                
            # Second column: Meal type
            worksheet.cell(row=row_idx, column=2, value=meal_type)
            
            # Third column: Total
            total = day_data.get("Total", 0)
            worksheet.cell(row=row_idx, column=3, value=total if total > 0 else "")
            
            # Day columns
            for day in range(1, 32):
                day_str = str(day)
                count = day_data.get(day_str, 0)
                worksheet.cell(row=row_idx, column=3+day, value=count if count > 0 else "")
            
            # Apply borders to all cells in the row
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            row_idx += 1
            first_meal = False
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_time_report(workbook, year, month, month_number, output_dir, prompt_for_location):
    """Generate a time-based monthly report with three worksheets"""
    try:
        # Create worksheets
        daily_worksheet = workbook.active
        daily_worksheet.title = "Daily Summary"
        user_worksheet = workbook.create_sheet("User Summary")
        detail_worksheet = workbook.create_sheet("Consumption Detail")
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
            
        cursor = conn.cursor(dictionary=True)
        
        # Get meal schedule (time slots) - Ideally from settings, but we'll query distinct values
        meal_schedule_query = f"""
            SELECT DISTINCT 
                CASE 
                    WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                    WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                    ELSE 'Dinner'
                END as MealType
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
            ORDER BY 
                CASE 
                    WHEN HOUR(PunchDateTime) < 10 THEN 1
                    WHEN HOUR(PunchDateTime) < 15 THEN 2
                    ELSE 3
                END
        """
        cursor.execute(meal_schedule_query, (year, month))
        meal_types = [row['MealType'] for row in cursor.fetchall()]
        
        # If no meal types found, use defaults
        if not meal_types:
            meal_types = ["Breakfast", "Lunch", "Dinner"]
        
        # Fill Daily Summary worksheet
        fill_time_daily_summary(daily_worksheet, cursor, year, month, meal_types)
        
        # Fill User Summary worksheet
        fill_time_user_summary(user_worksheet, cursor, year, month, meal_types)
        
        # Fill Consumption Detail worksheet
        fill_time_consumption_detail(detail_worksheet, cursor, year, month, meal_types)
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Create the file path
        file_name = f"{year}_{month_number}_CanteenReports.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Ask for save location if prompted
        if prompt_for_location:
            # This would typically use a GUI dialog in a desktop app
            # For a script, we'll just use the output_dir
            pass
            
        # Save the workbook
        workbook.save(file_path)
        print(f"Time-based monthly report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating time report: {e}")
        import traceback
        traceback.print_exc()
        return None

def fill_time_daily_summary(worksheet, cursor, year, month, meal_types):
    """Fill the Daily Summary worksheet for time-based report"""
    # Define headers
    columns = ["Date"] + meal_types + ["Total"]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get daily counts by time-based meal type
    query = f"""
        SELECT 
            DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END as MealType,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY DATE_FORMAT(PunchDateTime, '%d-%m-%Y'), 
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END
        ORDER BY DATE(PunchDateTime)
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by date
    data_by_date = {}
    for row in results:
        date = row['PunchDate']
        meal = row['MealType']
        count = row['Count']
        
        if date not in data_by_date:
            data_by_date[date] = {meal_type: 0 for meal_type in meal_types}
            data_by_date[date]['Date'] = date
            data_by_date[date]['Total'] = 0
            
        data_by_date[date][meal] = count
        data_by_date[date]['Total'] += count
    
    # Add data rows
    row_idx = 2
    for date, data in sorted(data_by_date.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_idx > 1:  # Center align all but first column
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {meal_type: 0 for meal_type in meal_types}
    totals['Date'] = 'Total'
    totals['Total'] = 0
    
    for date_data in data_by_date.values():
        for meal_type in meal_types:
            totals[meal_type] += date_data.get(meal_type, 0)
        totals['Total'] += date_data.get('Total', 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, 0)
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_idx > 1:  # Center align all but first column
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def fill_time_user_summary(worksheet, cursor, year, month, meal_types):
    """Fill the User Summary worksheet for time-based report"""
    # Define headers
    columns = ["Year", "Month", "PunchID", "Name"] + meal_types
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get user data
    query = f"""
        SELECT 
            PunchCardNo,
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END as MealType,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, 
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END
        ORDER BY PunchCardNo
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user
    data_by_user = {}
    for row in results:
        user_id = row['PunchCardNo']
        meal = row['MealType']
        count = row['Count']
        
        if user_id not in data_by_user:
            data_by_user[user_id] = {meal_type: 0 for meal_type in meal_types}
            data_by_user[user_id]['Year'] = year
            data_by_user[user_id]['Month'] = month
            data_by_user[user_id]['PunchID'] = user_id
            data_by_user[user_id]['Name'] = ""  # No name available in this table
            
        data_by_user[user_id][meal] = count
    
    # Add data rows
    row_idx = 2
    for user_id, data in sorted(data_by_user.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0 if col_name in meal_types else "")
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_name in meal_types:  # Center align meal counts
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {meal_type: 0 for meal_type in meal_types}
    totals['PunchID'] = 'Total'
    
    for user_data in data_by_user.values():
        for meal_type in meal_types:
            totals[meal_type] += user_data.get(meal_type, 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, "")
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_name in meal_types:  # Center align meal counts
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def fill_time_consumption_detail(worksheet, cursor, year, month, meal_types):
    """Fill the Consumption Detail worksheet for time-based report"""
    # Define headers - User, Date, Total, then day numbers 1-31
    columns = ["User", "Date", "Total"] + [str(i) for i in range(1, 32)]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get detailed consumption data
    query = f"""
        SELECT 
            PunchCardNo,
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END as MealType,
            DAY(PunchDateTime) as Day,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, 
            CASE 
                WHEN HOUR(PunchDateTime) < 10 THEN 'Breakfast'
                WHEN HOUR(PunchDateTime) < 15 THEN 'Lunch'
                ELSE 'Dinner'
            END, 
            DAY(PunchDateTime)
        ORDER BY PunchCardNo, MealType, Day
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user and meal type
    data_by_user_meal = {}
    
    for row in results:
        user_id = row['PunchCardNo']
        meal_type = row['MealType']
        day = row['Day']
        count = row['Count']
        
        if user_id not in data_by_user_meal:
            data_by_user_meal[user_id] = {}
            
        if meal_type not in data_by_user_meal[user_id]:
            data_by_user_meal[user_id][meal_type] = {"Total": 0}
            
        data_by_user_meal[user_id][meal_type][str(day)] = count
        data_by_user_meal[user_id][meal_type]["Total"] += count
    
    # Add data rows
    row_idx = 2
    for user_id, meal_data in sorted(data_by_user_meal.items()):
        first_meal = True
        for meal_type in meal_types:
            if meal_type not in meal_data:
                continue
                
            day_data = meal_data[meal_type]
                
            # First column: User ID (only for first meal type)
            if first_meal:
                cell = worksheet.cell(row=row_idx, column=1, value=f"{user_id}")
                cell.font = Font(bold=True)
            else:
                worksheet.cell(row=row_idx, column=1, value="")
                
            # Second column: Meal type
            worksheet.cell(row=row_idx, column=2, value=meal_type)
            
            # Third column: Total
            total = day_data.get("Total", 0)
            worksheet.cell(row=row_idx, column=3, value=total if total > 0 else "")
            
            # Day columns
            for day in range(1, 32):
                day_str = str(day)
                count = day_data.get(day_str, 0)
                worksheet.cell(row=row_idx, column=3+day, value=count if count > 0 else "")
            
            # Apply borders to all cells in the row
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            row_idx += 1
            first_meal = False
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_timebase_monthly_report(year, month, report_type=None, output_dir="Reports/monthly", prompt_for_location=True):
    """
    Generate a monthly report similar to the timeBase.js format with three sheets:
    - Daily Summary: Shows counts by day
    - User Summary: Shows counts by user
    - Consumption Detail: Shows detailed consumption data by user and day
    
    Args:
        year (int): Year for the report
        month (int): Month for the report (1-12)
        report_type (str): Type of report - 'deviceoptions', 'timeoptions', or 'menuoptions'
        output_dir (str): Directory to save the output file
        prompt_for_location (bool): Whether to prompt for file save location
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate inputs
        if not isinstance(year, int) or not isinstance(month, int):
            raise ValueError("Year and month must be integers")
        
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
            
        # Set month name and number
        month_names = ["January", "February", "March", "April", "May", "June", 
                      "July", "August", "September", "October", "November", "December"]
        month_name = month_names[month-1]
        month_number = f"{month:02d}"
        
        # Get canteen type if not provided
        if not report_type:
            # Try to load from settings (can be extended later)
            report_type = "deviceoptions"  # Default
            
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        # Create workbook
        workbook = Workbook()
        
        # Generate the appropriate report based on type
        report_type = report_type.lower().strip()
        if report_type == "deviceoptions":
            file_path = generate_device_report(workbook, year, month, month_number, output_dir, prompt_for_location)
        elif report_type == "timeoptions":
            file_path = generate_time_report(workbook, year, month, month_number, output_dir, prompt_for_location)
        elif report_type == "menuoptions":
            file_path = generate_menu_report(workbook, year, month, month_number, output_dir, prompt_for_location)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
            
        return file_path
        
    except Exception as e:
        print(f"Error generating timebase monthly report: {e}")
        import traceback
        traceback.print_exc()
        return None

def generate_menu_report(workbook, year, month, month_number, output_dir, prompt_for_location):
    """Generate a menu-based monthly report with three worksheets"""
    try:
        # Create worksheets
        daily_worksheet = workbook.active
        daily_worksheet.title = "Daily Summary"
        user_worksheet = workbook.create_sheet("User Summary")
        detail_worksheet = workbook.create_sheet("Consumption Detail")
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
            
        cursor = conn.cursor(dictionary=True)
        
        # Get menu items (food items)
        menu_items_query = f"""
            SELECT DISTINCT Fooditem 
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
            ORDER BY Fooditem
        """
        cursor.execute(menu_items_query, (year, month))
        menu_items = [row['Fooditem'] for row in cursor.fetchall()]
        
        # Fill Daily Summary worksheet
        fill_menu_daily_summary(daily_worksheet, cursor, year, month, menu_items)
        
        # Fill User Summary worksheet
        fill_menu_user_summary(user_worksheet, cursor, year, month, menu_items)
        
        # Fill Consumption Detail worksheet
        fill_menu_consumption_detail(detail_worksheet, cursor, year, month, menu_items)
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Create the file path
        file_name = f"{year}_{month_number}_CanteenReports.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Ask for save location if prompted
        if prompt_for_location:
            # This would typically use a GUI dialog in a desktop app
            # For a script, we'll just use the output_dir
            pass
            
        # Save the workbook
        workbook.save(file_path)
        print(f"Menu-based monthly report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating menu report: {e}")
        import traceback
        traceback.print_exc()
        return None

def fill_menu_daily_summary(worksheet, cursor, year, month, menu_items):
    """Fill the Daily Summary worksheet for menu-based report"""
    # Define headers
    columns = ["Date"] + menu_items + ["Total"]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get daily counts by menu item
    query = f"""
        SELECT 
            DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
            Fooditem,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY DATE_FORMAT(PunchDateTime, '%d-%m-%Y'), Fooditem
        ORDER BY DATE(PunchDateTime)
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by date
    data_by_date = {}
    for row in results:
        date = row['PunchDate']
        food_item = row['Fooditem']
        count = row['Count']
        
        if date not in data_by_date:
            data_by_date[date] = {item: 0 for item in menu_items}
            data_by_date[date]['Date'] = date
            data_by_date[date]['Total'] = 0
            
        data_by_date[date][food_item] = count
        data_by_date[date]['Total'] += count
    
    # Add data rows
    row_idx = 2
    for date, data in sorted(data_by_date.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_idx > 1:  # Center align all but first column
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {item: 0 for item in menu_items}
    totals['Date'] = 'Total'
    totals['Total'] = 0
    
    for date_data in data_by_date.values():
        for item in menu_items:
            totals[item] += date_data.get(item, 0)
        totals['Total'] += date_data.get('Total', 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, 0)
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_idx > 1:  # Center align all but first column
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def fill_menu_user_summary(worksheet, cursor, year, month, menu_items):
    """Fill the User Summary worksheet for menu-based report"""
    # Define headers
    columns = ["Year", "Month", "PunchID", "Name"] + menu_items
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get user data
    query = f"""
        SELECT 
            PunchCardNo,
            Fooditem,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, Fooditem
        ORDER BY PunchCardNo
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user
    data_by_user = {}
    for row in results:
        user_id = row['PunchCardNo']
        food_item = row['Fooditem']
        count = row['Count']
        
        if user_id not in data_by_user:
            data_by_user[user_id] = {item: 0 for item in menu_items}
            data_by_user[user_id]['Year'] = year
            data_by_user[user_id]['Month'] = month
            data_by_user[user_id]['PunchID'] = user_id
            data_by_user[user_id]['Name'] = ""  # No name available in this table
            
        data_by_user[user_id][food_item] = count
    
    # Add data rows
    row_idx = 2
    for user_id, data in sorted(data_by_user.items()):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, 0 if col_name in menu_items else "")
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            if col_name in menu_items:  # Center align food item counts
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
    
    # Add totals row
    totals = {item: 0 for item in menu_items}
    totals['PunchID'] = 'Total'
    
    for user_data in data_by_user.values():
        for item in menu_items:
            totals[item] += user_data.get(item, 0)
    
    for col_idx, col_name in enumerate(columns, 1):
        value = totals.get(col_name, "")
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        if col_name in menu_items:  # Center align food item counts
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def fill_menu_consumption_detail(worksheet, cursor, year, month, menu_items):
    """Fill the Consumption Detail worksheet for menu-based report"""
    # Define headers - Employee, Date, Total, then day numbers 1-31
    columns = ["Employee", "Date", "Total"] + [str(i) for i in range(1, 32)]
    
    # Add header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Query to get detailed consumption data
    query = f"""
        SELECT 
            PunchCardNo,
            Fooditem,
            DAY(PunchDateTime) as Day,
            COUNT(*) as Count
        FROM {DB_TABLE}
        WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        GROUP BY PunchCardNo, Fooditem, DAY(PunchDateTime)
        ORDER BY PunchCardNo, Fooditem, Day
    """
    
    cursor.execute(query, (year, month))
    results = cursor.fetchall()
    
    # Process data by user and food item
    data_by_user_item = {}
    
    for row in results:
        user_id = row['PunchCardNo']
        food_item = row['Fooditem']
        day = row['Day']
        count = row['Count']
        
        if user_id not in data_by_user_item:
            data_by_user_item[user_id] = {}
            
        if food_item not in data_by_user_item[user_id]:
            data_by_user_item[user_id][food_item] = {"Total": 0}
            
        data_by_user_item[user_id][food_item][str(day)] = count
        data_by_user_item[user_id][food_item]["Total"] += count
    
    # Add data rows
    row_idx = 2
    for user_id, item_data in sorted(data_by_user_item.items()):
        first_item = True
        for food_item in menu_items:
            if food_item not in item_data:
                continue
                
            day_data = item_data[food_item]
                
            # First column: User ID (only for first food item)
            if first_item:
                cell = worksheet.cell(row=row_idx, column=1, value=f"{user_id}")
                cell.font = Font(bold=True)
            else:
                worksheet.cell(row=row_idx, column=1, value="")
                
            # Second column: Food item
            worksheet.cell(row=row_idx, column=2, value=food_item)
            
            # Third column: Total
            total = day_data.get("Total", 0)
            worksheet.cell(row=row_idx, column=3, value=total if total > 0 else "")
            
            # Day columns
            for day in range(1, 32):
                day_str = str(day)
                count = day_data.get(day_str, 0)
                worksheet.cell(row=row_idx, column=3+day, value=count if count > 0 else "")
            
            # Apply borders to all cells in the row
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            row_idx += 1
            first_item = False
    
    # Auto-adjust column width
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

# Example usage
if __name__ == "__main__":
    # Generate traditional reports for May 2025
    # generate_monthly_report(2025, 5)
    generate_daily_report("2025-05-22")
    generate_logs_report(2025, 5)
    generate_fooditem_count_report()
    
    # Generate timebase reports with different formats
    print("\n=== Generating TimeBase Format Reports ===")
    # Device-based report
    generate_timebase_monthly_report(2025, 5, "deviceoptions", output_dir="Reports/monthly")
    # Time-based report
    generate_timebase_monthly_report(2025, 5, "timeoptions", output_dir="Reports/monthly")
    # Menu-based report
    generate_timebase_monthly_report(2025, 5, "menuoptions", output_dir="Reports/monthly") 
