import os
import pandas as pd
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Database configuration
DB_HOST = "103.216.211.36"
DB_USER = "pgcanteen"
DB_PORT = 33975
DB_PASS = "L^{Z,8~zzfF9(nd8"
DB_NAME = "payguru_canteen"
DB_TABLE = "sequentiallog"

def connect_to_database():
    """
    Create and return a connection to the database
    """
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            port=DB_PORT,
            password=DB_PASS,
            database=DB_NAME
        )
        return conn
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

def generate_monthly_report(year, month, output_dir="Reports/monthly"):
    """
    Generate a monthly report showing count of meal types per day

    Args:
        year (int): Year for the report
        month (int): Month for the report (1-12)
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate inputs
        if not isinstance(year, int) or not isinstance(month, int):
            raise ValueError("Year and month must be integers")
        
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Get distinct meal types
        meal_types_query = f"""
            SELECT DISTINCT Fooditem 
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
        """
        cursor.execute(meal_types_query, (year, month))
        meal_types = [row['Fooditem'] for row in cursor.fetchall()]
        
        # Query to get meal counts by date for the specified month
        query = f"""
            SELECT 
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                Fooditem,
                COUNT(*) as Count
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = %s AND MONTH(PunchDateTime) = %s
            GROUP BY DATE_FORMAT(PunchDateTime, '%d-%m-%Y'), Fooditem
            ORDER BY DATE(PunchDateTime)
        """
        
        cursor.execute(query, (year, month))
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {month}/{year}")
            # Create empty dataframe with dynamic columns
            columns = ["Date"] + meal_types + ["Total"]
            df = pd.DataFrame(columns=columns)
        else:
            # Convert to DataFrame for easier manipulation
            data_df = pd.DataFrame(results)
            
            # Create a pivot table
            pivot_df = pd.pivot_table(
                data_df,
                values="Count",
                index="PunchDate",
                columns="Fooditem",
                aggfunc='sum',
                fill_value=0
            )
            
            # Ensure all meal types from query are present
            for meal_type in meal_types:
                if meal_type not in pivot_df.columns:
                    pivot_df[meal_type] = 0
            
            # Keep only the required columns and reorder them
            pivot_df = pivot_df[meal_types]
            
            # Add Total column
            pivot_df["Total"] = pivot_df.sum(axis=1)
            
            # Reset index to make PunchDate a column
            pivot_df = pivot_df.reset_index()
            
            # Rename column to match expected format
            pivot_df = pivot_df.rename(columns={"PunchDate": "Date"})
            
            # Calculate totals
            totals = pivot_df[meal_types + ["Total"]].sum().to_frame().T
            totals.insert(0, "Date", "Total")
            
            # Combine dataframe with totals
            df = pd.concat([pivot_df, totals], ignore_index=True)
        
        # Create Excel file
        file_name = f"{year}_{month:02d}_CanteenReports.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Monthly Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Monthly Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align except for the first column
                    if col_num > 1:
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Format total row with bold
                    if row_num == len(df) + 1:
                        cell.font = Font(bold=True)
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Monthly report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating monthly report: {e}")
        return None
def generate_daily_report(date, output_dir="Reports/daily"):
    """
    Generate a daily report showing all entries for a specific date
    
    Args:
        date (str): Date in format YYYY-MM-DD
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate date format
        datetime.strptime(date, "%Y-%m-%d")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to get all entries for the specified date
        query = f"""
            SELECT 
                PunchCardNo,
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                TIME_FORMAT(PunchDateTime, '%H:%i:%s') as PunchTime,
                Fooditem
            FROM {DB_TABLE}
            WHERE DATE(PunchDateTime) = '{date}'
            ORDER BY PunchDateTime
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {date}")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Punch ID",  "Punch Date", "Punch Time", "Meal Type"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Format the data
            df["Punch ID"] = df["PunchCardNo"]
            df["Meal Type"] = df["Fooditem"]
            
            # Select and reorder columns
            df = df[["Punch ID", "PunchDate", "PunchTime", "Meal Type"]]
            
            # Rename columns to match expected format
            df = df.rename(columns={"PunchDate": "Punch Date", "PunchTime": "Punch Time"})
        
        # Create Excel file
        formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%Y-%m-%d")
        file_name = f"{formatted_date}_DailyConsumption.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Daily Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Daily Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align time and meal type columns
                    if col_num in [4, 5]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Daily report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating daily report: {e}")
        return None

def generate_logs_report(year, month, output_dir="Reports/CanteenLogs"):
    """
    Generate a logs report showing all entries for a specific month
    
    Args:
        year (int): Year for the report
        month (int): Month for the report (1-12)
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Validate inputs
        if not isinstance(year, int) or not isinstance(month, int):
            raise ValueError("Year and month must be integers")
        
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to get all entries for the specified month
        query = f"""
            SELECT 
                PunchCardNo,
                DATE_FORMAT(PunchDateTime, '%d-%m-%Y') as PunchDate,
                TIME_FORMAT(PunchDateTime, '%H:%i:%s') as PunchTime,
                Fooditem
            FROM {DB_TABLE}
            WHERE YEAR(PunchDateTime) = {year} AND MONTH(PunchDateTime) = {month}
            ORDER BY PunchDateTime
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print(f"No data found for {month}/{year}")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Punch ID", "Punch Date", "Punch Time", "Meal Type"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Format the data
            # Since we can't join with configh for name, just use PunchCardNo
            df["Punch ID"] = df["PunchCardNo"]
            df["Meal Type"] = df["Fooditem"]
            
            # Select and reorder columns - use directly formatted columns from SQL
            df = df[["Punch ID", "PunchDate", "PunchTime", "Meal Type"]]
            
            # Rename columns to match expected format
            df = df.rename(columns={"PunchDate": "Punch Date", "PunchTime": "Punch Time"})
        
        # Create Excel file
        file_name = f"{year}_{month}_CanteenLogs.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Logs Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Logs Report']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align time and meal type columns
                    if col_num in [4, 5]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Logs report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating logs report: {e}")
        return None

def generate_fooditem_count_report(output_dir="Reports/fooditemEvaluation"):
    """
    Generate a report showing counts of each food item in the database
    
    Args:
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated Excel file or None if error
    """
    try:
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Connect to database
        conn = connect_to_database()
        if not conn:
            return None
        
        cursor = conn.cursor(dictionary=True)
        
        # Query to count occurrences of each food item
        query = f"""
            SELECT 
                Fooditem,
                COUNT(*) as Count
            FROM {DB_TABLE}
            GROUP BY Fooditem
            ORDER BY COUNT(*) DESC
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Process data
        if not results:
            print("No food items found in the database")
            # Create empty dataframe with required columns
            df = pd.DataFrame(columns=["Food Item", "Count"])
        else:
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Rename columns to match expected format
            df = df.rename(columns={"Fooditem": "Food Item"})
        
        # Create Excel file
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_name = f"{current_date}_FoodItemCounts.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Food Item Counts', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Food Item Counts']
            
            # Apply formatting
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Format all cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Center align count column
                    if col_num == 2:
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Food item count report generated: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"Error generating food item count report: {e}")
        return None

# Example usage
if __name__ == "__main__":
    # Generate reports for May 2025
    generate_monthly_report(2025, 5)
    generate_daily_report("2025-05-22")
    generate_logs_report(2025, 5)
    generate_fooditem_count_report() 